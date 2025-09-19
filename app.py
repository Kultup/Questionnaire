from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response, render_template_string
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from dotenv import load_dotenv
import os
import sys
import qrcode
import io
import base64
import secrets
import requests
from datetime import datetime, timedelta
from sqlalchemy import func, and_
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import csv
from io import StringIO
import logging
from logging.handlers import RotatingFileHandler
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Load environment variables
load_dotenv()

# Configure logging for production
def configure_logging(app):
    if not app.debug and not app.testing:
        # Create logs directory if it doesn't exist
        log_dir = os.path.dirname(os.getenv('LOG_FILE', 'logs/app.log'))
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        # Configure file handler
        file_handler = RotatingFileHandler(
            os.getenv('LOG_FILE', 'logs/app.log'),
            maxBytes=int(os.getenv('LOG_MAX_BYTES', 10485760)),
            backupCount=int(os.getenv('LOG_BACKUP_COUNT', 5))
        )
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(getattr(logging, os.getenv('LOG_LEVEL', 'INFO')))
        app.logger.addHandler(file_handler)
        app.logger.setLevel(getattr(logging, os.getenv('LOG_LEVEL', 'INFO')))
        app.logger.info('Feedback System startup')

app = Flask(__name__)

# Configuration
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')

# Use absolute path for SQLite database
basedir = os.path.abspath(os.path.dirname(__file__))
db_file_path = os.path.join(basedir, 'instance', 'feedback_system.db')
# Ensure proper URI format for Windows - use pathlib for better cross-platform support
from pathlib import Path
db_path_obj = Path(db_file_path)
default_db_path = f"sqlite:///{db_path_obj.as_posix()}"
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', default_db_path)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Disable CSRF protection
app.config['WTF_CSRF_ENABLED'] = os.getenv('WTF_CSRF_ENABLED', 'True').lower() == 'true'

# Email configuration
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS', 'True').lower() == 'true'
app.config['MAIL_USE_SSL'] = os.getenv('MAIL_USE_SSL', 'False').lower() == 'true'
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER')
app.config['MAIL_SUBJECT_PREFIX'] = os.getenv('MAIL_SUBJECT_PREFIX', '[Відгук] ')
app.config['MAIL_ENABLED'] = os.getenv('MAIL_ENABLED', 'False').lower() == 'true'

# Security settings for production
if os.getenv('FLASK_ENV') == 'production':
    app.config['SESSION_COOKIE_SECURE'] = os.getenv('SESSION_COOKIE_SECURE', 'True').lower() == 'true'
    app.config['SESSION_COOKIE_HTTPONLY'] = os.getenv('SESSION_COOKIE_HTTPONLY', 'True').lower() == 'true'
    app.config['SESSION_COOKIE_SAMESITE'] = os.getenv('SESSION_COOKIE_SAMESITE', 'Lax')
    app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 16777216))
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = int(os.getenv('SEND_FILE_MAX_AGE_DEFAULT', 31536000))

# Configure logging
configure_logging(app)

# Domain configuration from environment variables
DOMAIN = os.getenv('DOMAIN', 'localhost')
BASE_URL = os.getenv('BASE_URL', 'http://localhost:5000')
PORT = int(os.getenv('PORT', 5000))

# Helper function to generate external URLs with custom domain
def get_external_url(endpoint, **values):
    """Generate external URL using BASE_URL from environment variables"""
    if BASE_URL and BASE_URL != f'http://localhost:{PORT}':
        # Use custom domain from environment
        with app.test_request_context():
            path = url_for(endpoint, **values)
        return BASE_URL.rstrip('/') + path
    else:
        # Use Flask's default _external=True for localhost
        return url_for(endpoint, _external=True, **values)

# Import models and forms first
from models import db, User, Question, Survey, Answer, QRCode, Admin
from forms import LoginForm, UserForm, QuestionForm, SurveyForm, DateFilterForm, BotTokenForm
from telegram_service import TelegramService, create_feedback_message

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Будь ласка, увійдіть для доступу до цієї сторінки.'
login_manager.login_message_category = 'info'

def init_database():
    """Initialize the database with tables"""
    try:
        # Debug information
        print(f"🔍 Database URI: {app.config['SQLALCHEMY_DATABASE_URI']}")
        print(f"🔍 Database file path: {db_file_path}")
        
        # Ensure instance directory exists
        instance_dir = os.path.dirname(db_file_path)
        if not os.path.exists(instance_dir):
            os.makedirs(instance_dir)
            print(f"📁 Створено директорію: {instance_dir}")
        else:
            print(f"📁 Директорія вже існує: {instance_dir}")
        
        # Check directory permissions
        import stat
        dir_stat = os.stat(instance_dir)
        print(f"🔐 Права доступу до директорії: {oct(dir_stat.st_mode)}")
        
        with app.app_context():
            # Create all tables
            db.create_all()
            print("✅ Таблиці бази даних створено успішно!")
            
            # Check if database file was created
            if os.path.exists(db_file_path):
                file_size = os.path.getsize(db_file_path)
                print(f"📊 База даних: {db_file_path}")
                print(f"📏 Розмір файлу: {file_size} байт")
            else:
                print("⚠️ Файл бази даних не було створено!")
                
    except Exception as e:
        print(f"❌ Помилка ініціалізації бази даних: {e}")
        import traceback
        print(f"🔍 Детальна інформація про помилку:")
        traceback.print_exc()
        return False
    
    return True

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Helper functions
# Authorization decorators
def admin_required(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Доступ заборонено. Потрібна авторизація адміністратора.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def manager_required(f):
    """Decorator to require manager authentication and prevent admin access"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if admin is trying to access manager routes
        if 'admin_logged_in' in session:
            flash('Доступ заборонено. Адміністратор не може використовувати менеджерські функції.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Check if manager is authenticated
        if not current_user.is_authenticated:
            flash('Доступ заборонено. Потрібна авторизація менеджера.', 'error')
            return redirect(url_for('login'))
        
        return f(*args, **kwargs)
    return decorated_function

def validate_bot_token(bot_token):
    """Validate Telegram bot token by checking getMe API"""
    if not bot_token:
        return False
    
    try:
        url = f"https://api.telegram.org/bot{bot_token}/getMe"
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            return data.get('ok', False)
        return False
    except Exception as e:
        print(f"Bot token validation error: {e}")
        return False

def get_bot_info(bot_token):
    """Отримує інформацію про бота (username, ім'я)"""
    try:
        url = f"https://api.telegram.org/bot{bot_token}/getMe"
        response = requests.get(url, timeout=10)
        data = response.json()
        
        if data.get('ok'):
            bot_info = data['result']
            return {
                'username': bot_info.get('username'),
                'first_name': bot_info.get('first_name'),
                'id': bot_info.get('id')
            }
        return None
    except Exception as e:
        print(f"DEBUG: Помилка при отриманні інформації про бота: {e}")
        return None

def send_test_telegram_message(bot_token):
    """Send test message to Telegram bot to verify configuration"""
    if not bot_token:
        print("DEBUG: Токен бота не вказано")
        return False, "Токен бота не вказано"
    
    # Validate token before sending message
    if not validate_bot_token(bot_token):
        print("DEBUG: Невірний токен бота")
        return False, "Невірний токен бота"
    
    # Try to get updates to find a chat
    try:
        # Get bot updates to find available chats
        updates_url = f"https://api.telegram.org/bot{bot_token}/getUpdates"
        print(f"DEBUG: Отримую оновлення з {updates_url}")
        updates_response = requests.get(updates_url, timeout=10)
        print(f"DEBUG: Статус відповіді getUpdates: {updates_response.status_code}")
        
        if updates_response.status_code == 200:
            updates_data = updates_response.json()
            print(f"DEBUG: Дані оновлень: {updates_data}")
            
            if updates_data.get('ok') and updates_data.get('result'):
                # Get the most recent chat_id from updates
                chat_id = None
                for update in reversed(updates_data['result']):
                    if 'message' in update and 'chat' in update['message']:
                        chat_id = update['message']['chat']['id']
                        print(f"DEBUG: Знайдено chat_id: {chat_id}")
                        break
                
                if not chat_id:
                    print("DEBUG: Не знайдено чатів у оновленнях")
                    return False, "Не знайдено чатів. Спочатку надішліть повідомлення боту (наприклад, /start)."
            else:
                print("DEBUG: Немає результатів у відповіді або помилка API")
                return False, "Не знайдено повідомлень. Спочатку надішліть повідомлення боту (наприклад, /start)."
        else:
            print(f"DEBUG: Помилка HTTP при getUpdates: {updates_response.status_code}")
            return False, f"Помилка отримання оновлень бота: {updates_response.status_code}"
    except Exception as e:
        print(f"DEBUG: Виняток при getUpdates: {str(e)}")
        return False, f"Помилка з'єднання з Telegram: {str(e)}"
    
    # Send test message
    try:
        test_message = "🤖 <b>Тестове повідомлення</b>\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n✅ <b>Вітаємо!</b> Ваш Telegram бот успішно налаштований!\n\n📢 <b>Що буде далі:</b>\n• Ви отримуватимете сповіщення про нові відгуки\n• Повідомлення будуть красиво відформатовані\n• Кожен відгук містить детальну інформацію\n\n🎉 <i>Система готова до роботи!</i>"
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': test_message,
            'parse_mode': 'HTML'
        }
        print(f"DEBUG: Відправляю повідомлення до chat_id {chat_id}")
        response = requests.post(url, data=data, timeout=10)
        print(f"DEBUG: Статус відповіді sendMessage: {response.status_code}")
        print(f"DEBUG: Відповідь sendMessage: {response.text}")
        
        if response.status_code == 200:
            return True, "Тестове повідомлення успішно надіслано!"
        else:
            return False, f"Помилка відправки повідомлення: {response.status_code} - {response.text}"
    except Exception as e:
        print(f"DEBUG: Виняток при sendMessage: {str(e)}")
        return False, f"Помилка відправки: {str(e)}"

def send_telegram_message(bot_token, message, chat_id=None):
    """Send message to Telegram bot"""
    if not bot_token:
        error_msg = "No bot token provided"
        print(error_msg)
        return False, error_msg
    
    # Validate token before sending message
    if not validate_bot_token(bot_token):
        error_msg = f"Invalid bot token: {bot_token[:10]}..."
        print(error_msg)
        return False, error_msg
    
    # If no chat_id provided, try to get updates to find a chat
    if not chat_id:
        try:
            # Get bot updates to find available chats
            updates_url = f"https://api.telegram.org/bot{bot_token}/getUpdates"
            updates_response = requests.get(updates_url, timeout=10)
            if updates_response.status_code == 200:
                updates_data = updates_response.json()
                if updates_data.get('ok') and updates_data.get('result'):
                    # Get the most recent chat_id from updates
                    for update in reversed(updates_data['result']):
                        if 'message' in update and 'chat' in update['message']:
                            chat_id = update['message']['chat']['id']
                            break
                    
                    if not chat_id:
                        error_msg = "No chat_id found in bot updates. Please send a message to the bot first."
                        print(error_msg)
                        return False, error_msg
                else:
                    error_msg = "No updates found for bot. Please send a message to the bot first."
                    print(error_msg)
                    return False, error_msg
            else:
                error_msg = f"Failed to get bot updates: {updates_response.status_code}"
                print(error_msg)
                return False, error_msg
        except Exception as e:
            error_msg = f"Error getting bot updates: {e}"
            print(error_msg)
            return False, error_msg
    
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'HTML'
        }
        response = requests.post(url, data=data, timeout=10)
        
        if response.status_code == 200:
            success_msg = f"Message sent successfully to chat_id: {chat_id}"
            print(success_msg)
            return True, success_msg
        else:
            error_msg = f"Failed to send message. Status: {response.status_code}, Response: {response.text}"
            print(error_msg)
            return False, error_msg
    except Exception as e:
        error_msg = f"Telegram error: {e}"
        print(error_msg)
        return False, error_msg

def send_email_message(email_address, subject, message):
    """Send email message to specified address"""
    if not app.config['MAIL_ENABLED']:
        print("Email sending is disabled")
        return False
    
    if not email_address:
        print("No email address provided")
        return False
    
    if not app.config['MAIL_USERNAME'] or not app.config['MAIL_PASSWORD']:
        print("Email credentials not configured")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"{app.config['MAIL_SUBJECT_PREFIX']}{subject}"
        msg['From'] = app.config['MAIL_DEFAULT_SENDER'] or app.config['MAIL_USERNAME']
        msg['To'] = email_address
        
        # Create HTML and plain text versions
        text_message = message.replace('<b>', '').replace('</b>', '').replace('<i>', '').replace('</i>', '').replace('<br>', '\n').replace('<br/>', '\n')
        # Remove emojis and special characters for plain text
        text_message = text_message.replace('🍽️', '').replace('👨‍💼', '').replace('⭐', '').replace('━', '-').replace('❓', '').replace('✅', '[ТАК]').replace('❌', '[НІ]').replace('💬', '').replace('📝', '').replace('💭', '')
        
        # Create beautiful HTML version
        html_message = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 10px 10px 0 0; text-align: center; }}
                .content {{ background: #f9f9f9; padding: 20px; border-radius: 0 0 10px 10px; border: 1px solid #ddd; }}
                .question {{ background: white; margin: 15px 0; padding: 15px; border-radius: 8px; border-left: 4px solid #667eea; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                .answer {{ margin: 10px 0; padding: 10px; background: #e8f4fd; border-radius: 5px; }}
                .comment {{ margin: 10px 0; padding: 10px; background: #fff3cd; border-radius: 5px; font-style: italic; }}
                .separator {{ border-top: 2px solid #667eea; margin: 20px 0; }}
                .rating {{ font-size: 18px; font-weight: bold; color: #ff6b35; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>{message.split('<b>')[1].split('</b>')[0] if '<b>' in message else 'Новий відгук'}</h2>
            </div>
            <div class="content">
        """
        
        # Process message content for HTML
        lines = message.split('\n')
        in_question = False
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '👨‍💼' in line or '⭐' in line:
                html_message += f'<p style="margin: 10px 0; font-weight: bold;">{line}</p>'
            elif '━' in line:
                html_message += '<div class="separator"></div>'
            elif '❓' in line or '📝' in line:
                if in_question:
                    html_message += '</div>'
                html_message += f'<div class="question"><h4 style="margin: 0 0 10px 0; color: #667eea;">{line}</h4>'
                in_question = True
            elif '✅' in line or '❌' in line or '💭' in line:
                html_message += f'<div class="answer">{line}</div>'
            elif '💬' in line:
                html_message += f'<div class="comment">{line}</div>'
            else:
                html_message += f'<p>{line}</p>'
        
        if in_question:
            html_message += '</div>'
            
        html_message += """
            </div>
        </body>
        </html>
        """
        
        # Attach parts
        part1 = MIMEText(text_message, 'plain', 'utf-8')
        part2 = MIMEText(html_message, 'html', 'utf-8')
        
        msg.attach(part1)
        msg.attach(part2)
        
        # Send email
        if app.config['MAIL_USE_SSL']:
            server = smtplib.SMTP_SSL(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        else:
            server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
            if app.config['MAIL_USE_TLS']:
                server.starttls()
        
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        server.send_message(msg)
        server.quit()
        
        print(f"Email sent successfully to: {email_address}")
        return True
        
    except Exception as e:
        print(f"Email error: {e}")
        return False

@app.route('/admin/export/<int:user_id>')
@admin_required
def admin_export_user(user_id):
    
    # Get the specific user
    user = User.query.get_or_404(user_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Відгуки - {user.restaurant_name}"
    
    # Headers
    headers = ['Дата', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Get surveys for this specific user
    surveys = Survey.query.filter_by(user_id=user_id).all()
    
    for survey in surveys:
        answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
        
        if answers:
            for answer in answers:
                ws.append([
                    survey.created_at.strftime('%Y-%m-%d %H:%M'),
                    survey.waiter_name,
                    survey.overall_score,
                    answer.question.question_text,
                    'Так' if answer.answer else 'Ні',
                    answer.comment or ''
                ])
        else:
            ws.append([
                survey.created_at.strftime('%Y-%m-%d %H:%M'),
                survey.waiter_name,
                survey.overall_score,
                '', '', ''
            ])
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={user.restaurant_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def generate_qr_code(data):
    """Generate QR code and return binary PNG data"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Return binary PNG data
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer.getvalue()

# Middleware for route protection
@app.before_request
def check_route_access():
    """Middleware to automatically check access to protected routes"""
    # Skip checks for static files and public routes
    if request.endpoint in ['static', 'index', 'survey', 'survey_submit', 'survey_success', 'login', 'logout']:
        return
    
    # Admin routes protection
    if request.endpoint and (request.endpoint.startswith('admin') or '/admin/' in request.path):
        if 'admin_logged_in' not in session:
            flash('Доступ заборонено. Потрібна авторизація адміністратора.', 'error')
            return redirect(url_for('login'))
        
        # Prevent managers from accessing admin routes (only if they are logged in as managers)
        if current_user.is_authenticated and 'admin_logged_in' not in session:
            flash('Доступ заборонено. Менеджер не може використовувати адміністративні функції.', 'error')
            return redirect(url_for('user_dashboard'))
    
    # Manager routes protection  
    if request.endpoint and (request.endpoint.startswith('user') or '/user/' in request.path):
        # Prevent admin from accessing manager routes
        if 'admin_logged_in' in session:
            flash('Доступ заборонено. Адміністратор не може використовувати менеджерські функції.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Check manager authentication
        if not current_user.is_authenticated:
            flash('Доступ заборонено. Потрібна авторизація менеджера.', 'error')
            return redirect(url_for('login'))

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        # Check if admin (from environment variables)
        admin_login = os.getenv('ADMIN_LOGIN', 'admin')
        admin_password = os.getenv('ADMIN_PASSWORD', 'admin123')
        
        if form.login.data == admin_login and form.password.data == admin_password:
            session['admin_logged_in'] = True
            session['admin_id'] = 1  # Static admin ID
            flash('Успішний вхід як адміністратор!', 'success')
            return redirect(url_for('admin_dashboard'))
        
        # Check if user
        user = User.query.filter_by(login=form.login.data).first()
        if user and user.is_active and check_password_hash(user.password_hash, form.password.data):
            login_user(user)
            flash(f'Ласкаво просимо, {user.restaurant_name}!', 'success')
            return redirect(url_for('user_dashboard'))
        
        flash('Невірний логін або пароль', 'error')
    
    return render_template('login.html', form=form)

@app.route('/logout')
def logout():
    if 'admin_logged_in' in session:
        session.pop('admin_logged_in', None)
        session.pop('admin_id', None)
        flash('Ви вийшли з системи', 'info')
    else:
        logout_user()
        flash('Ви вийшли з системи', 'info')
    return redirect(url_for('index'))

# Admin routes
@app.route('/admin')
@admin_required
def admin_dashboard():
    # Get statistics
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    total_surveys = Survey.query.count()
    
    # Calculate average score across all surveys
    average_score_result = db.session.query(func.avg(Survey.overall_score)).scalar()
    average_score = round(average_score_result, 1) if average_score_result else 0
    
    # Today's surveys
    today = datetime.now().date()
    today_surveys = Survey.query.filter(func.date(Survey.created_at) == today).count()
    
    # Recent surveys
    recent_surveys = Survey.query.order_by(Survey.created_at.desc()).limit(10).all()
    
    # Users with survey counts
    users_with_counts = db.session.query(
        User,
        func.count(Survey.id).label('survey_count')
    ).outerjoin(Survey).group_by(User.id).all()
    
    # Convert to list of users with survey_count attribute
    users = []
    for user, survey_count in users_with_counts:
        user.survey_count = survey_count
        users.append(user)
    
    # Chart data for surveys over time (last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    surveys_by_date = db.session.query(
        func.date(Survey.created_at).label('date'),
        func.count(Survey.id).label('count')
    ).filter(
        Survey.created_at >= thirty_days_ago
    ).group_by(func.date(Survey.created_at)).all()
    
    # Prepare chart data
    chart_labels = []
    chart_data = []
    
    # Fill in missing dates with 0 counts
    current_date = thirty_days_ago.date()
    end_date = datetime.now().date()
    
    # Convert string dates to date objects for proper comparison
    surveys_dict = {}
    for item in surveys_by_date:
        if isinstance(item.date, str):
            # SQLite returns dates as strings, convert to date object
            date_obj = datetime.strptime(item.date, '%Y-%m-%d').date()
        else:
            date_obj = item.date
        surveys_dict[date_obj] = item.count
    
    while current_date <= end_date:
        chart_labels.append(current_date.strftime('%m-%d'))
        chart_data.append(surveys_dict.get(current_date, 0))
        current_date += timedelta(days=1)
    
    # Ratings distribution
    ratings_distribution = [0, 0, 0, 0]  # [1-3, 4-6, 7-8, 9-10]
    all_surveys = Survey.query.all()
    
    for survey in all_surveys:
        score = survey.overall_score
        if 1 <= score <= 3:
            ratings_distribution[0] += 1
        elif 4 <= score <= 6:
            ratings_distribution[1] += 1
        elif 7 <= score <= 8:
            ratings_distribution[2] += 1
        elif 9 <= score <= 10:
            ratings_distribution[3] += 1
    
    return render_template('admin/dashboard.html',
                         total_users=total_users,
                         active_users=active_users,
                         total_surveys=total_surveys,
                         average_score=average_score,
                         today_surveys=today_surveys,
                         recent_surveys=recent_surveys,
                         users=users,
                         chart_labels=chart_labels,
                         chart_data=chart_data,
                         ratings_distribution=ratings_distribution)

@app.route('/admin/users')
@admin_required
def admin_users():
    
    users = User.query.all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/new', methods=['GET', 'POST'])
@admin_required
def admin_add_user():
    
    form = UserForm()
    if form.validate_on_submit():
        # Generate unique token
        unique_token = secrets.token_urlsafe(16)
        while User.query.filter_by(unique_token=unique_token).first():
            unique_token = secrets.token_urlsafe(16)
        
        user = User(
            login=form.login.data,
            password_hash=generate_password_hash(form.password.data),
            restaurant_name=form.restaurant_name.data,
            city=form.city.data,
            unique_token=unique_token,
            bot_token=form.bot_token.data,
            email_address=form.email_address.data,
            email_enabled=form.email_enabled.data,
            is_active=True
        )
        
        db.session.add(user)
        db.session.commit()
        
        # Send test message if bot_token is provided
        if form.bot_token.data:
            success, message = send_test_telegram_message(form.bot_token.data)
            if success:
                flash(f'Менеджер {user.restaurant_name} успішно створений! {message}', 'success')
            else:
                flash(f'Менеджер {user.restaurant_name} створений, але помилка з ботом: {message}', 'warning')
        else:
            flash(f'Менеджер {user.restaurant_name} успішно створений!', 'success')
        
        return redirect(url_for('admin_users'))
    
    return render_template('admin/user_form.html', form=form, user=None)

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_user(user_id):
    
    user = User.query.get_or_404(user_id)
    form = UserForm(original_user=user, obj=user)
    
    if form.validate_on_submit():
        # Store old bot_token to check if it changed
        old_bot_token = user.bot_token
        
        user.login = form.login.data
        if form.password.data:
            user.password_hash = generate_password_hash(form.password.data)
        user.restaurant_name = form.restaurant_name.data
        user.city = form.city.data
        user.bot_token = form.bot_token.data
        user.email_address = form.email_address.data
        user.email_enabled = form.email_enabled.data
        user.is_active = form.is_active.data
        
        db.session.commit()
        
        # Send test message if bot_token was added or changed
        if form.bot_token.data and form.bot_token.data != old_bot_token:
            success, message = send_test_telegram_message(form.bot_token.data)
            if success:
                flash(f'Менеджер {user.restaurant_name} оновлений! {message}', 'success')
            else:
                flash(f'Менеджер {user.restaurant_name} оновлений, але помилка з ботом: {message}', 'warning')
        else:
            flash(f'Менеджер {user.restaurant_name} оновлений!', 'success')
        
        return redirect(url_for('admin_users'))
    
    return render_template('admin/user_form.html', form=form, user=user)

@app.route('/admin/users/<int:user_id>')
@admin_required
def admin_user_details(user_id):
    
    user = User.query.get_or_404(user_id)
    
    # Get user statistics
    total_surveys = Survey.query.filter_by(user_id=user.id).count()
    avg_score = db.session.query(func.avg(Survey.overall_score)).filter_by(user_id=user.id).scalar() or 0
    
    # Recent surveys
    recent_surveys = Survey.query.filter_by(user_id=user.id).order_by(Survey.created_at.desc()).limit(5).all()
    
    # Questions
    questions = Question.query.filter_by(user_id=user.id).all()
    
    # Generate chart data for this specific user (last 30 days)
    period_ago = datetime.now() - timedelta(days=30)
    
    # Get surveys by date for this user
    surveys_by_date = db.session.query(
        func.date(Survey.created_at).label('date'),
        func.count(Survey.id).label('count')
    ).filter(
        Survey.created_at >= period_ago,
        Survey.user_id == user.id
    ).group_by(func.date(Survey.created_at)).all()
    
    # Prepare chart data
    chart_labels = []
    chart_data = []
    
    # Fill in missing dates with 0 counts
    current_date = period_ago.date()
    end_date = datetime.now().date()
    
    # Convert string dates to date objects for proper comparison
    surveys_dict = {}
    for item in surveys_by_date:
        if isinstance(item.date, str):
            # SQLite returns dates as strings, convert to date object
            date_obj = datetime.strptime(item.date, '%Y-%m-%d').date()
        else:
            date_obj = item.date
        surveys_dict[date_obj] = item.count
    
    while current_date <= end_date:
        chart_labels.append(current_date.strftime('%m-%d'))
        chart_data.append(surveys_dict.get(current_date, 0))
        current_date += timedelta(days=1)
    
    # Get ratings distribution for this user
    ratings_distribution = [0, 0, 0, 0]  # [1-3, 4-6, 7-8, 9-10]
    user_surveys = Survey.query.filter(
        Survey.created_at >= period_ago,
        Survey.user_id == user.id
    ).all()
    
    for survey in user_surveys:
        score = survey.overall_score
        if 1 <= score <= 3:
            ratings_distribution[0] += 1
        elif 4 <= score <= 6:
            ratings_distribution[1] += 1
        elif 7 <= score <= 8:
            ratings_distribution[2] += 1
        elif 9 <= score <= 10:
            ratings_distribution[3] += 1
    
    return render_template('admin/user_details.html',
                         user=user,
                         total_surveys=total_surveys,
                         avg_score=avg_score,
                         recent_surveys=recent_surveys,
                         questions=questions,
                         chart_labels=chart_labels,
                         chart_data=chart_data,
                         ratings_distribution=ratings_distribution)

@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_user(user_id):
    
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'активований' if user.is_active else 'деактивований'
    flash(f'Менеджер {user.restaurant_name} {status}!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    
    user = User.query.get_or_404(user_id)
    restaurant_name = user.restaurant_name
    
    # Delete related data - fix for SQLAlchemy join() + delete() issue
    # First get survey IDs for this user
    survey_ids = [s.id for s in Survey.query.filter_by(user_id=user_id).all()]
    
    # Delete answers for these surveys
    if survey_ids:
        Answer.query.filter(Answer.survey_id.in_(survey_ids)).delete(synchronize_session=False)
    
    # Delete other related data
    Survey.query.filter_by(user_id=user_id).delete()
    Question.query.filter_by(user_id=user_id).delete()
    QRCode.query.filter_by(user_id=user_id).delete()
    
    db.session.delete(user)
    db.session.commit()
    
    flash(f'Менеджер {restaurant_name} видалений!', 'success')
    return redirect(url_for('admin_users'))

# User routes
@app.route('/user')
@manager_required
def user_dashboard():
    # Get statistics
    total_surveys = Survey.query.filter_by(user_id=current_user.id).count()
    avg_score = db.session.query(func.avg(Survey.overall_score)).filter_by(user_id=current_user.id).scalar() or 0
    
    # Today's surveys
    today = datetime.now().date()
    today_surveys = Survey.query.filter(
        Survey.user_id == current_user.id,
        func.date(Survey.created_at) == today
    ).count()
    
    # Recent surveys
    recent_surveys = Survey.query.filter_by(user_id=current_user.id).order_by(Survey.created_at.desc()).limit(5).all()
    
    # Questions count
    total_questions = Question.query.filter_by(user_id=current_user.id).count()
    active_questions = Question.query.filter_by(user_id=current_user.id, is_active=True).count()
    
    # Chart data for surveys over time (last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    surveys_by_date = db.session.query(
        func.date(Survey.created_at).label('date'),
        func.count(Survey.id).label('count')
    ).filter(
        Survey.user_id == current_user.id,
        Survey.created_at >= thirty_days_ago
    ).group_by(func.date(Survey.created_at)).all()
    
    # Prepare chart data
    chart_labels = []
    chart_data = []
    
    # Fill in missing dates with 0 counts
    current_date = thirty_days_ago.date()
    end_date = datetime.now().date()
    surveys_dict = {item.date: item.count for item in surveys_by_date}
    
    while current_date <= end_date:
        chart_labels.append(current_date.strftime('%m-%d'))
        chart_data.append(surveys_dict.get(current_date, 0))
        current_date += timedelta(days=1)
    
    # Ratings distribution
    ratings_distribution = [0, 0, 0, 0]  # [1-3, 4-6, 7-8, 9-10]
    all_surveys = Survey.query.filter_by(user_id=current_user.id).all()
    
    for survey in all_surveys:
        score = survey.overall_score
        if 1 <= score <= 3:
            ratings_distribution[0] += 1
        elif 4 <= score <= 6:
            ratings_distribution[1] += 1
        elif 7 <= score <= 8:
            ratings_distribution[2] += 1
        elif 9 <= score <= 10:
            ratings_distribution[3] += 1
    
    # Recent comments
    recent_comments = db.session.query(Answer).join(Survey).filter(
        Survey.user_id == current_user.id,
        Answer.comment.isnot(None),
        Answer.comment != ''
    ).order_by(Survey.created_at.desc()).limit(5).all()
    
    # Questions statistics
    questions_stats = []
    user_questions = Question.query.filter_by(user_id=current_user.id, is_active=True).all()
    
    for question in user_questions:
        # Count answers for this question
        answers = Answer.query.join(Survey).filter(
            Survey.user_id == current_user.id,
            Answer.question_id == question.id
        ).all()
        
        if answers:
            yes_count = sum(1 for answer in answers if answer.answer)
            no_count = len(answers) - yes_count
            total_answers = len(answers)
            
            yes_percentage = (yes_count / total_answers * 100) if total_answers > 0 else 0
            no_percentage = (no_count / total_answers * 100) if total_answers > 0 else 0
            
            # Count comments for this question
            comments_count = sum(1 for answer in answers if answer.comment and answer.comment.strip())
            
            questions_stats.append({
                'question_text': question.question_text,
                'yes_count': yes_count,
                'no_count': no_count,
                'yes_percentage': yes_percentage,
                'no_percentage': no_percentage,
                'comments_count': comments_count
            })
    
    return render_template('user/dashboard.html',
                         total_surveys=total_surveys,
                         avg_score=avg_score,
                         today_surveys=today_surveys,
                         recent_surveys=recent_surveys,
                         total_questions=total_questions,
                         active_questions=active_questions,
                         chart_labels=chart_labels,
                         chart_data=chart_data,
                         ratings_distribution=ratings_distribution,
                         recent_comments=recent_comments,
                         questions_stats=questions_stats)

@app.route('/bot-instructions')
@login_required
def bot_instructions():
    """Сторінка з інструкціями по налаштуванню Telegram бота - доступна всім користувачам"""
    return render_template('bot_instructions.html')

@app.route('/user/questions')
@manager_required
def user_questions():
    questions = Question.query.filter_by(user_id=current_user.id).all()
    active_count = Question.query.filter_by(user_id=current_user.id, is_active=True).count()
    inactive_count = Question.query.filter_by(user_id=current_user.id, is_active=False).count()
    
    return render_template('user/questions.html',
                         questions=questions,
                         active_count=active_count,
                         inactive_count=inactive_count)

@app.route('/user/questions/add', methods=['POST'])
@manager_required
def user_add_question():
    form = QuestionForm()
    if form.validate_on_submit():
        question = Question(
            user_id=current_user.id,
            question_text=form.question_text.data,
            question_type=form.question_type.data,
            is_active=form.is_active.data
        )
        db.session.add(question)
        db.session.commit()
        flash('Питання додано успішно!', 'success')
    else:
        # Детальна обробка помилок валідації
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Помилка в полі "{field}": {error}', 'error')
        if not form.errors:
            flash('Помилка при додаванні питання', 'error')
    
    return redirect(url_for('user_questions'))

@app.route('/user/questions/<int:question_id>/edit', methods=['POST'])
@manager_required
def user_edit_question(question_id):
    question = Question.query.filter_by(id=question_id, user_id=current_user.id).first_or_404()
    form = QuestionForm()
    
    if form.validate_on_submit():
        question.question_text = form.question_text.data
        question.question_type = form.question_type.data
        question.is_active = form.is_active.data
        db.session.commit()
        flash('Питання оновлено успішно!', 'success')
    else:
        flash('Помилка при оновленні питання', 'error')
    
    return redirect(url_for('user_questions'))

@app.route('/user/questions/<int:question_id>/toggle', methods=['POST'])
@manager_required
def user_toggle_question(question_id):
    question = Question.query.filter_by(id=question_id, user_id=current_user.id).first_or_404()
    question.is_active = not question.is_active
    db.session.commit()
    
    status = 'активовано' if question.is_active else 'деактивовано'
    flash(f'Питання {status}!', 'success')
    return redirect(url_for('user_questions'))

@app.route('/user/questions/<int:question_id>/delete', methods=['POST'])
@manager_required
def user_delete_question(question_id):
    question = Question.query.filter_by(id=question_id, user_id=current_user.id).first_or_404()
    
    # Delete related answers
    Answer.query.filter_by(question_id=question_id).delete()
    
    db.session.delete(question)
    db.session.commit()
    
    flash('Питання видалено!', 'success')
    return redirect(url_for('user_questions'))

@app.route('/user/questions/bulk-delete', methods=['POST'])
@manager_required
def user_bulk_delete_questions():
    question_ids = request.form.getlist('question_ids')
    
    if not question_ids:
        flash('Не обрано жодного питання для видалення!', 'warning')
        return redirect(url_for('user_questions'))
    
    try:
        # Convert to integers and validate ownership
        question_ids = [int(qid) for qid in question_ids]
        
        # Get questions that belong to current user
        questions = Question.query.filter(
            Question.id.in_(question_ids),
            Question.user_id == current_user.id
        ).all()
        
        if not questions:
            flash('Не знайдено питань для видалення!', 'error')
            return redirect(url_for('user_questions'))
        
        deleted_count = 0
        for question in questions:
            # Delete related answers
            Answer.query.filter_by(question_id=question.id).delete()
            db.session.delete(question)
            deleted_count += 1
        
        db.session.commit()
        
        if deleted_count == 1:
            flash('Питання видалено!', 'success')
        else:
            flash(f'Видалено {deleted_count} питань!', 'success')
            
    except ValueError:
        flash('Помилка при обробці запиту!', 'error')
    except Exception as e:
        db.session.rollback()
        flash('Помилка при видаленні питань!', 'error')
    
    return redirect(url_for('user_questions'))

@app.route('/user/qr-code')
@manager_required
def user_qr_code():
    # Get or create QR code
    qr_code = QRCode.query.filter_by(user_id=current_user.id).first()
    
    if not qr_code:
        # Generate survey URL using dynamic domain configuration
        survey_url = get_external_url('survey', token=current_user.unique_token)
        qr_data = generate_qr_code(survey_url)
        
        qr_code = QRCode(
            user_id=current_user.id,
            qr_code_data=qr_data
        )
        db.session.add(qr_code)
        db.session.commit()
    
    # Get survey statistics
    total_surveys = Survey.query.filter_by(user_id=current_user.id).count()
    avg_score = db.session.query(func.avg(Survey.overall_score)).filter_by(user_id=current_user.id).scalar() or 0
    
    # Get active questions for preview
    questions = Question.query.filter_by(user_id=current_user.id, is_active=True).all()
    
    # Generate survey URL using dynamic domain configuration
    survey_url = get_external_url('survey', token=current_user.unique_token)
    
    # Convert binary QR code data to base64 for template display
    qr_code_base64 = base64.b64encode(qr_code.qr_code_data).decode() if qr_code else None
    
    # Count active questions
    active_questions_count = len(questions)
    
    return render_template('user/qr_code.html',
                         qr_code=qr_code,
                         qr_code_base64=qr_code_base64,
                         survey_url=survey_url,
                         total_surveys=total_surveys,
                         avg_score=avg_score,
                         average_score=avg_score,
                         questions=questions,
                         active_questions=questions,
                         active_questions_count=active_questions_count)

@app.route('/user/qr-code/regenerate', methods=['POST'])
@manager_required
def user_regenerate_qr():
    # Delete existing QR code
    QRCode.query.filter_by(user_id=current_user.id).delete()
    
    # Generate new unique token
    unique_token = secrets.token_urlsafe(16)
    while User.query.filter_by(unique_token=unique_token).first():
        unique_token = secrets.token_urlsafe(16)
    
    current_user.unique_token = unique_token
    db.session.commit()
    
    flash('QR-код перегенеровано!', 'success')
    return redirect(url_for('user_qr_code'))

# Public survey routes
@app.route('/survey/<token>', methods=['GET', 'POST'])
def survey(token):
    user = User.query.filter_by(unique_token=token).first_or_404()
    
    if not user.is_active:
        flash('Цей заклад тимчасово недоступний', 'error')
        return redirect(url_for('index'))
    
    questions = Question.query.filter_by(user_id=user.id, is_active=True).all()
    
    if request.method == 'POST':
        waiter_name = request.form.get('waiter_name')
        overall_score = request.form.get('overall_score')
        
        if not waiter_name or not overall_score:
            flash('Будь ласка, заповніть всі обов\'язкові поля', 'error')
            return render_template('survey.html', restaurant=user, questions=questions)
        
        # Create survey
        survey = Survey(
            user_id=user.id,
            waiter_name=waiter_name,
            overall_score=int(overall_score)
        )
        db.session.add(survey)
        db.session.flush()  # Get survey ID
        
        # Validate all required questions are answered
        validation_errors = []
        for i, question in enumerate(questions, 1):
            if question.question_type == 'yes_no':
                answer_value = request.form.get(f'question_{question.id}')
                if not answer_value:
                    # Truncate question text if too long or contains invalid characters
                    question_preview = question.question_text[:50] + "..." if len(question.question_text) > 50 else question.question_text
                    # Clean up invalid characters for display
                    question_preview = ''.join(c for c in question_preview if c.isprintable() and c not in ['ф', 'в', 'ц'] * 3)
                    if not question_preview.strip() or len(question_preview.strip()) < 3:
                        question_preview = f"питання №{i}"
                    validation_errors.append(f'Будь ласка, дайте відповідь на {question_preview}')
            else:
                # Manual input question
                comment = request.form.get(f'comment_{question.id}', '').strip()
                if not comment:
                    # Truncate question text if too long or contains invalid characters
                    question_preview = question.question_text[:50] + "..." if len(question.question_text) > 50 else question.question_text
                    # Clean up invalid characters for display
                    question_preview = ''.join(c for c in question_preview if c.isprintable() and c not in ['ф', 'в', 'ц'] * 3)
                    if not question_preview.strip() or len(question_preview.strip()) < 3:
                        question_preview = f"питання №{i}"
                    validation_errors.append(f'Будь ласка, дайте відповідь на {question_preview}')
        
        if validation_errors:
            for error in validation_errors:
                flash(error, 'error')
            return render_template('survey.html', restaurant=user, questions=questions)
        
        # Save answers
        telegram_message = f"🍽️ <b>Новий відгук для {user.restaurant_name}</b>\n\n"
        telegram_message += f"👨‍💼 <b>Офіціант:</b> {waiter_name}\n"
        telegram_message += f"⭐ <b>Загальна оцінка:</b> {overall_score}/10\n"
        telegram_message += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        
        for question in questions:
            comment = request.form.get(f'comment_{question.id}', '').strip()
            
            if question.question_type == 'yes_no':
                # Yes/No + Comment question
                answer_value = request.form.get(f'question_{question.id}')
                
                # We already validated that answer_value exists above
                answer = Answer(
                    survey_id=survey.id,
                    question_id=question.id,
                    answer=answer_value == 'true',
                    comment=comment if comment else None
                )
                db.session.add(answer)
                
                # Add to telegram message
                telegram_message += f"❓ <b>{question.question_text}</b>\n"
                answer_emoji = "✅" if answer_value == 'true' else "❌"
                telegram_message += f"{answer_emoji} <b>Відповідь:</b> {'Так' if answer_value == 'true' else 'Ні'}\n"
                if comment:
                    telegram_message += f"💬 <b>Коментар:</b> <i>{comment}</i>\n"
                telegram_message += "\n"
            else:
                # Manual input question - we already validated that comment exists above
                # For manual input questions, we set answer to False as a default since it's required
                answer = Answer(
                    survey_id=survey.id,
                    question_id=question.id,
                    answer=False,  # Default value for manual input questions
                    comment=comment
                )
                db.session.add(answer)
                
                # Add to telegram message
                telegram_message += f"📝 <b>{question.question_text}</b>\n"
                telegram_message += f"💭 <b>Відповідь:</b> <i>{comment}</i>\n\n"
        
        db.session.commit()
        
        # Send to Telegram group only
        if user.bot_token and user.telegram_group_enabled and user.telegram_group_id:
            telegram_service = TelegramService(user.bot_token)
            
            # Send to Telegram group
            group_result = telegram_service.send_message_to_chat(user.telegram_group_id, telegram_message)
            if not group_result['success']:
                print(f"Failed to send Telegram message to group for restaurant: {user.restaurant_name} (ID: {user.id}). Error: {group_result['error']}")
                # Log the error but don't show to survey user
        
        # Send to Email
        if user.email_enabled and user.email_address:
            subject = f"Новий відгук - {user.restaurant_name}"
            success = send_email_message(user.email_address, subject, telegram_message)
            if not success:
                print(f"Failed to send Email message for restaurant: {user.restaurant_name} (ID: {user.id})")
                # Note: We don't show error to survey user, but log it for admin
        
        # Store restaurant info in session for success page
        session['success_restaurant_id'] = user.id
        
        return redirect(url_for('survey_success'))
    
    return render_template('survey.html', restaurant=user, questions=questions)

@app.route('/survey/success')
def survey_success():
    restaurant_id = session.get('success_restaurant_id')
    if restaurant_id:
        restaurant = User.query.get(restaurant_id)
        # Clear the session data
        session.pop('success_restaurant_id', None)
        return render_template('survey_success.html', restaurant=restaurant)
    else:
        # Fallback if no restaurant info in session
        return render_template('survey_success.html')


@app.route('/api/user/chart-data')
@login_required
def user_chart_data():
    # Get period parameter (default to 30 days)
    period = request.args.get('period', '30', type=int)
    
    # Validate period
    if period not in [7, 30, 90]:
        period = 30
    
    # Get surveys by date for the specified period
    period_ago = datetime.now() - timedelta(days=period)
    
    surveys_by_date = db.session.query(
        func.date(Survey.created_at).label('date'),
        func.count(Survey.id).label('count')
    ).filter(
        Survey.user_id == current_user.id,
        Survey.created_at >= period_ago
    ).group_by(func.date(Survey.created_at)).all()
    
    # Prepare chart data with proper date range
    chart_labels = []
    chart_data = []
    
    # Fill in missing dates with 0 counts
    current_date = period_ago.date()
    end_date = datetime.now().date()
    
    # Convert string dates to date objects for proper comparison
    surveys_dict = {}
    for item in surveys_by_date:
        if isinstance(item.date, str):
            # SQLite returns dates as strings, convert to date object
            date_obj = datetime.strptime(item.date, '%Y-%m-%d').date()
        else:
            date_obj = item.date
        surveys_dict[date_obj] = item.count
    
    while current_date <= end_date:
        chart_labels.append(current_date.strftime('%m-%d'))
        chart_data.append(surveys_dict.get(current_date, 0))
        current_date += timedelta(days=1)
    
    # Get question statistics
    question_stats = []
    questions = Question.query.filter_by(user_id=current_user.id, is_active=True).all()
    
    for question in questions:
        yes_count = Answer.query.join(Survey).filter(
            Survey.user_id == current_user.id,
            Answer.question_id == question.id,
            Answer.answer == True
        ).count()
        
        no_count = Answer.query.join(Survey).filter(
            Survey.user_id == current_user.id,
            Answer.question_id == question.id,
            Answer.answer == False
        ).count()
        
        total_answers = yes_count + no_count
        
        question_stats.append({
            'question': question.question_text,
            'yes_count': yes_count,
            'no_count': no_count,
            'yes_percentage': round((yes_count / total_answers * 100) if total_answers > 0 else 0, 1),
            'no_percentage': round((no_count / total_answers * 100) if total_answers > 0 else 0, 1)
        })
    
    return jsonify({
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'question_stats': question_stats
    })

@app.route('/api/admin/chart-data')
@admin_required
def admin_chart_data():
    # Get period parameter (default to 30 days)
    period = request.args.get('period', '30', type=int)
    # Get user_id parameter for filtering by specific manager (optional)
    user_id = request.args.get('user_id', type=int)
    
    # Validate period
    if period not in [7, 30, 90]:
        period = 30
    
    # Get surveys by date for the specified period
    period_ago = datetime.now() - timedelta(days=period)
    
    # Build query with optional user filter
    query = db.session.query(
        func.date(Survey.created_at).label('date'),
        func.count(Survey.id).label('count')
    ).filter(Survey.created_at >= period_ago)
    
    # Add user filter if specified
    if user_id:
        query = query.filter(Survey.user_id == user_id)
    
    surveys_by_date = query.group_by(func.date(Survey.created_at)).all()
    
    # Prepare chart data with proper date range
    chart_labels = []
    chart_data = []
    
    # Fill in missing dates with 0 counts
    current_date = period_ago.date()
    end_date = datetime.now().date()
    surveys_dict = {item.date: item.count for item in surveys_by_date}
    
    while current_date <= end_date:
        chart_labels.append(current_date.strftime('%m-%d'))
        chart_data.append(surveys_dict.get(current_date, 0))
        current_date += timedelta(days=1)
    
    # Get ratings distribution
    ratings_distribution = [0, 0, 0, 0]  # [1-3, 4-6, 7-8, 9-10]
    
    # Build query for ratings with optional user filter
    surveys_query = Survey.query.filter(Survey.created_at >= period_ago)
    if user_id:
        surveys_query = surveys_query.filter(Survey.user_id == user_id)
    
    all_surveys = surveys_query.all()
    
    for survey in all_surveys:
        score = survey.overall_score
        if 1 <= score <= 3:
            ratings_distribution[0] += 1
        elif 4 <= score <= 6:
            ratings_distribution[1] += 1
        elif 7 <= score <= 8:
            ratings_distribution[2] += 1
        elif 9 <= score <= 10:
            ratings_distribution[3] += 1
    
    return jsonify({
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'ratings_distribution': ratings_distribution
    })

# Export routes
@app.route('/admin/export')
@admin_required
def admin_export():
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Всі відгуки"
        
        # Headers
        headers = ['Дата', 'Заклад', 'Місто', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Get all surveys with answers
        surveys = Survey.query.join(User).all()
        
        for survey in surveys:
            answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
            
            if answers:
                for answer in answers:
                    ws.append([
                        survey.created_at.strftime('%Y-%m-%d %H:%M'),
                        survey.user.restaurant_name,
                        survey.user.city,
                        survey.waiter_name,
                        survey.overall_score,
                        answer.question.question_text,
                        'Так' if answer.answer else 'Ні',
                        answer.comment or ''
                    ])
            else:
                ws.append([
                    survey.created_at.strftime('%Y-%m-%d %H:%M'),
                    survey.user.restaurant_name,
                    survey.user.city,
                    survey.waiter_name,
                    survey.overall_score,
                    '', '', ''
                ])
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        print(f"Export successful: {filename}")  # Debug log
        return response
        
    except Exception as e:
        print(f"Export error: {str(e)}")  # Debug log
        flash(f'Помилка при експорті: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/user/all-comments')
@manager_required
def user_all_comments():
    # Get all comments for the current user's surveys
    comments = db.session.query(Answer, Survey, Question).join(
        Survey, Answer.survey_id == Survey.id
    ).join(
        Question, Answer.question_id == Question.id
    ).filter(
        Survey.user_id == current_user.id,
        Answer.comment.isnot(None),
        Answer.comment != ''
    ).order_by(Survey.created_at.desc()).all()
    
    return render_template('user/all_comments.html', comments=comments)

@app.route('/user/all-surveys')
@manager_required
def user_all_surveys():
    # Get all surveys for the current user
    surveys = Survey.query.filter_by(user_id=current_user.id).order_by(Survey.created_at.desc()).all()
    
    # Get survey data with answers
    survey_data = []
    for survey in surveys:
        answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
        survey_data.append({
            'survey': survey,
            'answers': answers
        })
    
    return render_template('user/all_surveys.html', survey_data=survey_data)

@app.route('/user/filters', methods=['GET', 'POST'])
@manager_required
def user_filters():
    form = DateFilterForm()
    
    if form.validate_on_submit():
        # Redirect back to dashboard with filter parameters
        return redirect(url_for('user_dashboard', 
                               start_date=form.start_date.data, 
                               end_date=form.end_date.data))
    
    return render_template('user/filters.html', form=form)

@app.route('/user/email-settings', methods=['POST'])
@manager_required
def user_email_settings():
    """Handle email settings update for user"""
    email_address = request.form.get('email_address', '').strip()
    email_enabled = 'email_enabled' in request.form
    
    # Validate email if provided and enabled
    if email_enabled and not email_address:
        flash('Email адреса обов\'язкова, якщо увімкнені email сповіщення.', 'error')
        return redirect(url_for('user_bot_settings'))
    
    if email_address:
        # Basic email validation
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email_address):
            flash('Введіть коректну email адресу.', 'error')
            return redirect(url_for('user_bot_settings'))
    
    # Update user settings
    current_user.email_address = email_address if email_address else None
    current_user.email_enabled = email_enabled
    db.session.commit()
    
    if email_enabled and email_address:
        flash('Email налаштування успішно збережено. Ви будете отримувати копії відгуків на email.', 'success')
    elif email_address:
        flash('Email адреса збережена, але сповіщення вимкнені.', 'info')
    else:
        flash('Email сповіщення вимкнені.', 'info')
    
    return redirect(url_for('user_bot_settings'))


@app.route('/user/bot-settings', methods=['GET', 'POST'])
@manager_required
def user_bot_settings():
    form = BotTokenForm()
    bot_info = None
    
    # Get bot info if token exists
    if current_user.bot_token:
        bot_info = get_bot_info(current_user.bot_token)
    
    # Handle token removal
    if request.method == 'POST' and request.form.get('action') == 'remove_token':
        current_user.bot_token = None
        db.session.commit()
        flash('Токен бота успішно видалено.', 'success')
        return redirect(url_for('user_bot_settings'))
    
    # Handle form submission
    if form.validate_on_submit():
        # Simple validation for bot token format
        token = form.bot_token.data.strip()
        if not token:
            flash('Токен не може бути порожнім.', 'error')
        elif len(token) < 10:
            flash('Токен занадто короткий. Перевірте правильність токена.', 'error')
        else:
            # Store old token to check if it changed
            old_token = current_user.bot_token
            
            # Validate token with Telegram API
            if validate_bot_token(token):
                current_user.bot_token = token
                db.session.commit()
                
                # Send test message if token was added or changed
                print(f"DEBUG: Порівняння токенів - новий: '{token}', старий: '{old_token}'")
                if token != old_token:
                    print("DEBUG: Токен змінився, відправляю тестове повідомлення")
                    success, message = send_test_telegram_message(token)
                    if success:
                        flash(f'Токен бота успішно збережено та перевірено. {message}', 'success')
                    else:
                        # Get bot info for instructions
                        bot_info = get_bot_info(token)
                        if bot_info and bot_info.get('username'):
                            flash(f'Токен збережено, але для отримання сповіщень спочатку знайдіть бота @{bot_info["username"]} в Telegram і надішліть йому повідомлення /start', 'warning')
                        else:
                            flash(f'Токен збережено, але помилка з тестовим повідомленням: {message}', 'warning')
                else:
                    print("DEBUG: Токен не змінився, тестове повідомлення не відправляється")
                    flash('Токен бота успішно збережено та перевірено.', 'success')
                
                return redirect(url_for('user_bot_settings'))
            else:
                flash('Невірний токен бота. Перевірте правильність токена та спробуйте знову.', 'error')
    
    # Pre-fill form with current token
    if current_user.bot_token and not form.bot_token.data:
        form.bot_token.data = current_user.bot_token
    
    return render_template('user/bot_settings.html', form=form, bot_info=bot_info)

@app.route('/user/telegram-group-settings', methods=['POST'])
@manager_required
def user_telegram_group_settings():
    """Обробка налаштувань групи Telegram"""
    telegram_group_id = request.form.get('telegram_group_id', '').strip()
    telegram_enabled = request.form.get('telegram_enabled') == 'on'
    
    if telegram_enabled and not telegram_group_id:
        flash('Для увімкнення сповіщень необхідно вказати ID групи Telegram.', 'error')
        return redirect(url_for('user_bot_settings'))
    
    if telegram_enabled and not current_user.bot_token:
        flash('Для увімкнення сповіщень спочатку налаштуйте токен бота.', 'error')
        return redirect(url_for('user_bot_settings'))
    
    # Збереження налаштувань
    if telegram_enabled:
        current_user.telegram_group_id = telegram_group_id
        current_user.telegram_group_enabled = True
        flash('Налаштування групи Telegram збережено.', 'success')
    else:
        current_user.telegram_group_enabled = False
        # Зберігаємо ID групи, але вимикаємо сповіщення
        if telegram_group_id:
            current_user.telegram_group_id = telegram_group_id
        flash('Сповіщення в групу Telegram вимкнені.', 'info')
    
    db.session.commit()
    return redirect(url_for('user_bot_settings'))

@app.route('/user/test-group-message', methods=['POST'])
@manager_required
def user_test_group_message():
    """Відправка тестового повідомлення в групу Telegram"""
    if not current_user.bot_token:
        flash('Спочатку налаштуйте токен бота.', 'error')
        return redirect(url_for('user_bot_settings'))
    
    if not current_user.telegram_group_id:
        flash('Спочатку налаштуйте ID групи Telegram.', 'error')
        return redirect(url_for('user_bot_settings'))
    
    # Відправка тестового повідомлення
    test_message = f"🧪 Тестове повідомлення від {current_user.restaurant_name}\n\nЯкщо ви бачите це повідомлення, налаштування групи працюють правильно!"
    
    success, message = send_telegram_message(
        current_user.bot_token, 
        test_message, 
        current_user.telegram_group_id
    )
    
    if success:
        flash('Тестове повідомлення успішно відправлено в групу!', 'success')
    else:
        flash(f'Помилка відправки тестового повідомлення: {message}', 'error')
    
    return redirect(url_for('user_bot_settings'))

@app.route('/user/get-available-groups', methods=['POST'])
@manager_required
def user_get_available_groups():
    """Отримання списку доступних груп з оновлень бота"""
    if not current_user.bot_token:
        return jsonify({
            'success': False,
            'error': 'Спочатку налаштуйте токен бота.',
            'groups': []
        })
    
    try:
        telegram_service = TelegramService(current_user.bot_token)
        result = telegram_service.get_available_groups()
        
        if result['success']:
            return jsonify({
                'success': True,
                'groups': result['groups'],
                'error': None
            })
        else:
            return jsonify({
                'success': False,
                'error': result['error'],
                'groups': []
            })
            
    except Exception as e:
        app.logger.error(f"Error getting available groups for user {current_user.id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Виникла помилка при отриманні списку груп.',
            'groups': []
        })

@app.route('/user/export')
@manager_required
def user_export():
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Відгуки - {current_user.restaurant_name}"
    
    # Headers
    headers = ['Дата', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Get user surveys with answers
    surveys = Survey.query.filter_by(user_id=current_user.id).all()
    
    for survey in surveys:
        answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
        
        if answers:
            for answer in answers:
                ws.append([
                    survey.created_at.strftime('%Y-%m-%d %H:%M'),
                    survey.waiter_name,
                    survey.overall_score,
                    answer.question.question_text,
                    'Так' if answer.answer else 'Ні',
                    answer.comment or ''
                ])
        else:
            ws.append([
                survey.created_at.strftime('%Y-%m-%d %H:%M'),
                survey.waiter_name,
                survey.overall_score,
                '', '', ''
            ])
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={current_user.restaurant_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

# Enhanced export routes
@app.route('/user/export-page')
@manager_required
def user_export_page():
    """Сторінка з фільтрами для експорту (користувач)"""
    # Отримуємо унікальних офіціантів для поточного користувача
    waiters = db.session.query(Survey.waiter_name).filter_by(user_id=current_user.id).distinct().all()
    waiters = [w[0] for w in waiters if w[0]]
    
    return render_template('user/export.html', waiters=waiters)

@app.route('/user/export-filtered', methods=['POST'])
@manager_required
def user_export_filtered():
    """Фільтрований експорт для користувача"""
    try:
        # Отримуємо параметри фільтрів
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')
        waiters = request.form.getlist('waiters')
        rating_from = request.form.get('rating_from')
        rating_to = request.form.get('rating_to')
        only_comments = request.form.get('only_comments') == 'on'
        export_format = request.form.get('export_format', 'excel')
        include_summary = request.form.get('include_summary') == 'on'
        group_by_waiter = request.form.get('group_by_waiter') == 'on'
        
        # Базовий запит
        query = Survey.query.filter_by(user_id=current_user.id)
        
        # Застосовуємо фільтри
        if date_from:
            query = query.filter(Survey.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            date_to_end = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Survey.created_at < date_to_end)
        if waiters:
            query = query.filter(Survey.waiter_name.in_(waiters))
        if rating_from:
            query = query.filter(Survey.overall_score >= int(rating_from))
        if rating_to:
            query = query.filter(Survey.overall_score <= int(rating_to))
        
        surveys = query.all()
        
        if export_format == 'csv':
            return export_to_csv(surveys, current_user, only_comments, include_summary, group_by_waiter)
        else:
            return export_to_excel(surveys, current_user, only_comments, include_summary, group_by_waiter)
            
    except Exception as e:
        flash(f'Помилка при експорті: {str(e)}', 'error')
        return redirect(url_for('user_export_page'))

@app.route('/user/export-preview', methods=['POST'])
@manager_required
def user_export_preview():
    """Попередній перегляд даних для експорту (користувач)"""
    try:
        # Отримуємо параметри фільтрів (аналогічно до user_export_filtered)
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')
        waiters = request.form.getlist('waiters')
        rating_from = request.form.get('rating_from')
        rating_to = request.form.get('rating_to')
        only_comments = request.form.get('only_comments') == 'on'
        
        # Базовий запит
        query = Survey.query.filter_by(user_id=current_user.id)
        
        # Застосовуємо фільтри
        if date_from:
            query = query.filter(Survey.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            date_to_end = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Survey.created_at < date_to_end)
        if waiters:
            query = query.filter(Survey.waiter_name.in_(waiters))
        if rating_from:
            query = query.filter(Survey.overall_score >= int(rating_from))
        if rating_to:
            query = query.filter(Survey.overall_score <= int(rating_to))
        
        surveys = query.limit(10).all()  # Обмежуємо для попереднього перегляду
        
        # Підраховуємо статистику
        total_surveys = query.count()
        avg_rating = db.session.query(func.avg(Survey.overall_score)).filter(
            Survey.id.in_([s.id for s in query.all()])
        ).scalar() or 0
        
        preview_data = []
        for survey in surveys:
            answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
            if answers:
                for answer in answers:
                    if only_comments and not answer.comment:
                        continue
                    preview_data.append({
                        'date': survey.created_at.strftime('%Y-%m-%d %H:%M'),
                        'waiter': survey.waiter_name,
                        'rating': survey.overall_score,
                        'question': answer.question.question_text,
                        'answer': 'Так' if answer.answer else 'Ні',
                        'comment': answer.comment or ''
                    })
        
        return render_template_string('''
        <div class="alert alert-info">
            <strong>Статистика:</strong> Знайдено {{ total_surveys }} відгуків, середня оцінка: {{ "%.1f"|format(avg_rating) }}
        </div>
        <div class="table-responsive">
            <table class="table table-sm">
                <thead>
                    <tr>
                        <th>Дата</th>
                        <th>Офіціант</th>
                        <th>Оцінка</th>
                        <th>Питання</th>
                        <th>Відповідь</th>
                        <th>Коментар</th>
                    </tr>
                </thead>
                <tbody>
                    {% for row in preview_data %}
                    <tr>
                        <td>{{ row.date }}</td>
                        <td>{{ row.waiter }}</td>
                        <td>{{ row.rating }}</td>
                        <td>{{ row.question }}</td>
                        <td>{{ row.answer }}</td>
                        <td>{{ row.comment }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% if preview_data|length == 10 %}
        <div class="alert alert-warning">
            Показано перші 10 записів. Всього буде експортовано {{ total_surveys }} записів.
        </div>
        {% endif %}
        ''', preview_data=preview_data, total_surveys=total_surveys, avg_rating=avg_rating)
        
    except Exception as e:
        return f'<div class="alert alert-danger">Помилка: {str(e)}</div>'

@app.route('/admin/export-page')
@admin_required
def admin_export_page():
    """Сторінка з фільтрами для експорту (адмін)"""
    # Отримуємо дані для фільтрів
    restaurants = User.query.filter(User.role.in_(['manager', 'user'])).all()
    waiters = db.session.query(Survey.waiter_name).distinct().all()
    waiters = [w[0] for w in waiters if w[0]]
    cities = db.session.query(User.city).distinct().all()
    cities = [c[0] for c in cities if c[0]]
    
    return render_template('admin/export.html', 
                         restaurants=restaurants, 
                         waiters=waiters, 
                         cities=cities)

@app.route('/admin/export-filtered', methods=['POST'])
@admin_required
def admin_export_filtered():
    """Фільтрований експорт для адміністратора"""
    try:
        # Отримуємо параметри фільтрів
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')
        restaurants = request.form.getlist('restaurants')
        waiters = request.form.getlist('waiters')
        cities = request.form.getlist('cities')
        rating_from = request.form.get('rating_from')
        rating_to = request.form.get('rating_to')
        only_comments = request.form.get('only_comments') == 'on'
        export_format = request.form.get('export_format', 'excel')
        include_summary = request.form.get('include_summary') == 'on'
        group_by_restaurant = request.form.get('group_by_restaurant') == 'on'
        group_by_waiter = request.form.get('group_by_waiter') == 'on'
        
        # Базовий запит
        query = Survey.query.join(User)
        
        # Застосовуємо фільтри
        if date_from:
            query = query.filter(Survey.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            date_to_end = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Survey.created_at < date_to_end)
        if restaurants:
            query = query.filter(User.id.in_(restaurants))
        if waiters:
            query = query.filter(Survey.waiter_name.in_(waiters))
        if cities:
            query = query.filter(User.city.in_(cities))
        if rating_from:
            query = query.filter(Survey.overall_score >= int(rating_from))
        if rating_to:
            query = query.filter(Survey.overall_score <= int(rating_to))
        
        surveys = query.all()
        
        if export_format == 'csv':
            return export_to_csv_admin(surveys, only_comments, include_summary, group_by_restaurant, group_by_waiter)
        else:
            return export_to_excel_admin(surveys, only_comments, include_summary, group_by_restaurant, group_by_waiter)
            
    except Exception as e:
        flash(f'Помилка при експорті: {str(e)}', 'error')
        return redirect(url_for('admin_export_page'))

@app.route('/admin/export-preview', methods=['POST'])
@admin_required
def admin_export_preview():
    """Попередній перегляд даних для експорту (адмін)"""
    try:
        # Отримуємо параметри фільтрів (аналогічно до admin_export_filtered)
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')
        restaurants = request.form.getlist('restaurants')
        waiters = request.form.getlist('waiters')
        cities = request.form.getlist('cities')
        rating_from = request.form.get('rating_from')
        rating_to = request.form.get('rating_to')
        only_comments = request.form.get('only_comments') == 'on'
        
        # Базовий запит
        query = Survey.query.join(User)
        
        # Застосовуємо фільтри
        if date_from:
            query = query.filter(Survey.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            date_to_end = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Survey.created_at < date_to_end)
        if restaurants:
            query = query.filter(User.id.in_(restaurants))
        if waiters:
            query = query.filter(Survey.waiter_name.in_(waiters))
        if cities:
            query = query.filter(User.city.in_(cities))
        if rating_from:
            query = query.filter(Survey.overall_score >= int(rating_from))
        if rating_to:
            query = query.filter(Survey.overall_score <= int(rating_to))
        
        surveys = query.limit(10).all()  # Обмежуємо для попереднього перегляду
        
        # Підраховуємо статистику
        total_surveys = query.count()
        avg_rating = db.session.query(func.avg(Survey.overall_score)).filter(
            Survey.id.in_([s.id for s in query.all()])
        ).scalar() or 0
        
        preview_data = []
        for survey in surveys:
            answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
            if answers:
                for answer in answers:
                    if only_comments and not answer.comment:
                        continue
                    preview_data.append({
                        'date': survey.created_at.strftime('%Y-%m-%d %H:%M'),
                        'restaurant': survey.user.restaurant_name,
                        'city': survey.user.city,
                        'waiter': survey.waiter_name,
                        'rating': survey.overall_score,
                        'question': answer.question.question_text,
                        'answer': 'Так' if answer.answer else 'Ні',
                        'comment': answer.comment or ''
                    })
        
        return render_template_string('''
        <div class="alert alert-info">
            <strong>Статистика:</strong> Знайдено {{ total_surveys }} відгуків, середня оцінка: {{ "%.1f"|format(avg_rating) }}
        </div>
        <div class="table-responsive">
            <table class="table table-sm">
                <thead>
                    <tr>
                        <th>Дата</th>
                        <th>Заклад</th>
                        <th>Місто</th>
                        <th>Офіціант</th>
                        <th>Оцінка</th>
                        <th>Питання</th>
                        <th>Відповідь</th>
                        <th>Коментар</th>
                    </tr>
                </thead>
                <tbody>
                    {% for row in preview_data %}
                    <tr>
                        <td>{{ row.date }}</td>
                        <td>{{ row.restaurant }}</td>
                        <td>{{ row.city }}</td>
                        <td>{{ row.waiter }}</td>
                        <td>{{ row.rating }}</td>
                        <td>{{ row.question }}</td>
                        <td>{{ row.answer }}</td>
                        <td>{{ row.comment }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% if preview_data|length == 10 %}
        <div class="alert alert-warning">
            Показано перші 10 записів. Всього буде експортовано {{ total_surveys }} записів.
        </div>
        {% endif %}
        ''', preview_data=preview_data, total_surveys=total_surveys, avg_rating=avg_rating)
        
    except Exception as e:
        return f'<div class="alert alert-danger">Помилка: {str(e)}</div>'

# Enhanced export functions
def export_to_excel(surveys, user, only_comments=False, include_summary=False, group_by_waiter=False):
    """Покращений експорт в Excel з підтримкою фільтрів"""
    try:
        wb = openpyxl.Workbook()
        
        if group_by_waiter:
            # Групуємо по офіціантах
            waiters_data = {}
            for survey in surveys:
                waiter = survey.waiter_name or 'Невідомий офіціант'
                if waiter not in waiters_data:
                    waiters_data[waiter] = []
                waiters_data[waiter].append(survey)
            
            # Видаляємо стандартний лист
            wb.remove(wb.active)
            
            for waiter, waiter_surveys in waiters_data.items():
                ws = wb.create_sheet(title=waiter[:31])  # Excel обмежує назву листа до 31 символу
                _fill_worksheet(ws, waiter_surveys, only_comments, f"Звіт по офіціанту: {waiter}")
        else:
            ws = wb.active
            ws.title = "Експорт даних"
            _fill_worksheet(ws, surveys, only_comments, f"Звіт по закладу: {user.restaurant_name}")
        
        # Додаємо лист зі статистикою
        if include_summary:
            summary_ws = wb.create_sheet(title="Статистика")
            _fill_summary_worksheet(summary_ws, surveys)
        
        # Створюємо відповідь
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Помилка при створенні Excel файлу: {str(e)}', 'error')
        return redirect(url_for('user_export_page'))

def export_to_csv(surveys, user, only_comments=False, include_summary=False, group_by_waiter=False):
    """Експорт в CSV формат"""
    try:
        output = StringIO()
        writer = csv.writer(output)
        
        # Заголовки
        headers = ['Дата', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
        writer.writerow(headers)
        
        # Дані
        for survey in surveys:
            answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
            if answers:
                for answer in answers:
                    if only_comments and not answer.comment:
                        continue
                    writer.writerow([
                        survey.created_at.strftime('%Y-%m-%d %H:%M'),
                        survey.waiter_name or '',
                        survey.overall_score,
                        answer.question.question_text,
                        'Так' if answer.answer else 'Ні',
                        answer.comment or ''
                    ])
        
        # Створюємо відповідь
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        flash(f'Помилка при створенні CSV файлу: {str(e)}', 'error')
        return redirect(url_for('user_export_page'))

def export_to_excel_admin(surveys, only_comments=False, include_summary=False, group_by_restaurant=False, group_by_waiter=False):
    """Покращений експорт в Excel для адміністратора"""
    try:
        wb = openpyxl.Workbook()
        
        if group_by_restaurant:
            # Групуємо по закладах
            restaurants_data = {}
            for survey in surveys:
                restaurant = survey.user.restaurant_name or 'Невідомий заклад'
                if restaurant not in restaurants_data:
                    restaurants_data[restaurant] = []
                restaurants_data[restaurant].append(survey)
            
            # Видаляємо стандартний лист
            wb.remove(wb.active)
            
            for restaurant, restaurant_surveys in restaurants_data.items():
                ws = wb.create_sheet(title=restaurant[:31])
                _fill_worksheet_admin(ws, restaurant_surveys, only_comments, f"Звіт по закладу: {restaurant}")
                
        elif group_by_waiter:
            # Групуємо по офіціантах
            waiters_data = {}
            for survey in surveys:
                waiter = survey.waiter_name or 'Невідомий офіціант'
                if waiter not in waiters_data:
                    waiters_data[waiter] = []
                waiters_data[waiter].append(survey)
            
            # Видаляємо стандартний лист
            wb.remove(wb.active)
            
            for waiter, waiter_surveys in waiters_data.items():
                ws = wb.create_sheet(title=waiter[:31])
                _fill_worksheet_admin(ws, waiter_surveys, only_comments, f"Звіт по офіціанту: {waiter}")
        else:
            ws = wb.active
            ws.title = "Експорт даних"
            _fill_worksheet_admin(ws, surveys, only_comments, "Загальний звіт")
        
        # Додаємо лист зі статистикою
        if include_summary:
            summary_ws = wb.create_sheet(title="Статистика")
            _fill_summary_worksheet_admin(summary_ws, surveys)
        
        # Створюємо відповідь
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=admin_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Помилка при створенні Excel файлу: {str(e)}', 'error')
        return redirect(url_for('admin_export_page'))

def export_to_csv_admin(surveys, only_comments=False, include_summary=False, group_by_restaurant=False, group_by_waiter=False):
    """Експорт в CSV формат для адміністратора"""
    try:
        output = StringIO()
        writer = csv.writer(output)
        
        # Заголовки
        headers = ['Дата', 'Заклад', 'Місто', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
        writer.writerow(headers)
        
        # Дані
        for survey in surveys:
            answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
            if answers:
                for answer in answers:
                    if only_comments and not answer.comment:
                        continue
                    writer.writerow([
                        survey.created_at.strftime('%Y-%m-%d %H:%M'),
                        survey.user.restaurant_name or '',
                        survey.user.city or '',
                        survey.waiter_name or '',
                        survey.overall_score,
                        answer.question.question_text,
                        'Так' if answer.answer else 'Ні',
                        answer.comment or ''
                    ])
        
        # Створюємо відповідь
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=admin_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        flash(f'Помилка при створенні CSV файлу: {str(e)}', 'error')
        return redirect(url_for('admin_export_page'))

def _fill_worksheet(ws, surveys, only_comments, title):
    """Заповнює робочий лист даними для користувача"""
    # Заголовок
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Заголовки стовпців
    headers = ['Дата', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Дані
    row = 4
    for survey in surveys:
        answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
        if answers:
            for answer in answers:
                if only_comments and not answer.comment:
                    continue
                ws.cell(row=row, column=1, value=survey.created_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=2, value=survey.waiter_name or '')
                ws.cell(row=row, column=3, value=survey.overall_score)
                ws.cell(row=row, column=4, value=answer.question.question_text)
                ws.cell(row=row, column=5, value='Так' if answer.answer else 'Ні')
                ws.cell(row=row, column=6, value=answer.comment or '')
                row += 1
    
    # Автоширина стовпців
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def _fill_worksheet_admin(ws, surveys, only_comments, title):
    """Заповнює робочий лист даними для адміністратора"""
    # Заголовок
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Заголовки стовпців
    headers = ['Дата', 'Заклад', 'Місто', 'Офіціант', 'Загальна оцінка', 'Питання', 'Відповідь', 'Коментар']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Дані
    row = 4
    for survey in surveys:
        answers = Answer.query.filter_by(survey_id=survey.id).join(Question).all()
        if answers:
            for answer in answers:
                if only_comments and not answer.comment:
                    continue
                ws.cell(row=row, column=1, value=survey.created_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=2, value=survey.user.restaurant_name or '')
                ws.cell(row=row, column=3, value=survey.user.city or '')
                ws.cell(row=row, column=4, value=survey.waiter_name or '')
                ws.cell(row=row, column=5, value=survey.overall_score)
                ws.cell(row=row, column=6, value=answer.question.question_text)
                ws.cell(row=row, column=7, value='Так' if answer.answer else 'Ні')
                ws.cell(row=row, column=8, value=answer.comment or '')
                row += 1
    
    # Автоширина стовпців
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def _fill_summary_worksheet(ws, surveys):
    """Заповнює лист статистики для користувача"""
    ws.title = "Статистика"
    
    # Заголовок
    ws['A1'] = "Статистика відгуків"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Загальна статистика
    total_surveys = len(surveys)
    avg_rating = sum(s.overall_score for s in surveys) / total_surveys if total_surveys > 0 else 0
    
    ws['A3'] = "Загальна кількість відгуків:"
    ws['B3'] = total_surveys
    ws['A4'] = "Середня оцінка:"
    ws['B4'] = round(avg_rating, 2)
    
    # Статистика по офіціантах
    waiter_stats = {}
    for survey in surveys:
        waiter = survey.waiter_name or 'Невідомий'
        if waiter not in waiter_stats:
            waiter_stats[waiter] = {'count': 0, 'total_rating': 0}
        waiter_stats[waiter]['count'] += 1
        waiter_stats[waiter]['total_rating'] += survey.overall_score
    
    ws['A6'] = "Статистика по офіціантах:"
    ws['A6'].font = Font(bold=True)
    ws['A7'] = "Офіціант"
    ws['B7'] = "Кількість відгуків"
    ws['C7'] = "Середня оцінка"
    
    row = 8
    for waiter, stats in waiter_stats.items():
        ws.cell(row=row, column=1, value=waiter)
        ws.cell(row=row, column=2, value=stats['count'])
        ws.cell(row=row, column=3, value=round(stats['total_rating'] / stats['count'], 2))
        row += 1

def _fill_summary_worksheet_admin(ws, surveys):
    """Заповнює лист статистики для адміністратора"""
    ws.title = "Статистика"
    
    # Заголовок
    ws['A1'] = "Загальна статистика відгуків"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Загальна статистика
    total_surveys = len(surveys)
    avg_rating = sum(s.overall_score for s in surveys) / total_surveys if total_surveys > 0 else 0
    
    ws['A3'] = "Загальна кількість відгуків:"
    ws['B3'] = total_surveys
    ws['A4'] = "Середня оцінка:"
    ws['B4'] = round(avg_rating, 2)
    
    # Статистика по закладах
    restaurant_stats = {}
    for survey in surveys:
        restaurant = survey.user.restaurant_name or 'Невідомий заклад'
        if restaurant not in restaurant_stats:
            restaurant_stats[restaurant] = {'count': 0, 'total_rating': 0}
        restaurant_stats[restaurant]['count'] += 1
        restaurant_stats[restaurant]['total_rating'] += survey.overall_score
    
    ws['A6'] = "Статистика по закладах:"
    ws['A6'].font = Font(bold=True)
    ws['A7'] = "Заклад"
    ws['B7'] = "Кількість відгуків"
    ws['C7'] = "Середня оцінка"
    
    row = 8
    for restaurant, stats in restaurant_stats.items():
        ws.cell(row=row, column=1, value=restaurant)
        ws.cell(row=row, column=2, value=stats['count'])
        ws.cell(row=row, column=3, value=round(stats['total_rating'] / stats['count'], 2))
        row += 1

if __name__ == '__main__':
    print("🚀 Запуск Flask додатку...")
    
    # Автоматична ініціалізація бази даних
    print("🔧 Ініціалізація бази даних...")
    if init_database():
        print("🎉 База даних готова до роботи!")
    else:
        print("❌ Помилка ініціалізації бази даних!")
        sys.exit(1)
    
    # Production vs Development configuration
    is_production = os.getenv('FLASK_ENV') == 'production'
    
    if is_production:
        # Production settings
        app.run(
            host=os.getenv('HOST', '0.0.0.0'),
            port=int(os.getenv('PORT', 5000)),
            debug=False,
            threaded=True
        )
    else:
        # Development settings
        app.run(
            host=os.getenv('HOST', '127.0.0.1'),
            port=int(os.getenv('PORT', 5000)),
            debug=os.getenv('DEBUG', 'True').lower() == 'true'
        )